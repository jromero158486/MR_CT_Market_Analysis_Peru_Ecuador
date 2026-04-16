"""
build_excel_dashboard.py
------------------------
Genera el archivo Excel con:
1. Hoja de datos limpia (Power BI data source)
2. Tablas de análisis con formato profesional
3. Gráficos embebidos (para quien no tenga Power BI)
4. Hoja README con instrucciones para conectar Power BI
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart.label import DataLabelList
import os

# Paleta Siemens Healthineers
TEAL       = "00A0B0"   # Primary
DARK_TEAL  = "007080"
ORANGE     = "FF6B00"
LIGHT_GRAY = "F5F5F5"
MID_GRAY   = "CCCCCC"
DARK_GRAY  = "4A4A4A"
WHITE      = "FFFFFF"
BLACK      = "000000"

# ── Cargar datos ────────────────────────────────────────────────────────────
df = pd.read_excel("data/licitaciones_mr_ct.xlsx", sheet_name="Datos_Licitaciones")
df["fecha_convocatoria"] = pd.to_datetime(df["fecha_convocatoria"])

adj = df[df["estado"] == "Adjudicado"].copy()

# ── Tablas de análisis ──────────────────────────────────────────────────────
# Market share total
ms_total = (
    adj.groupby("marca_ganadora")
    .agg(licitaciones=("codigo", "count"),
         monto_total=("valor_referencial_USD", "sum"))
    .reset_index()
    .sort_values("licitaciones", ascending=False)
)
ms_total["share_%"] = (ms_total["licitaciones"] / ms_total["licitaciones"].sum() * 100).round(1)
ms_total["monto_MM_USD"] = (ms_total["monto_total"] / 1_000_000).round(2)

# Tendencia por año y modalidad
tend = (
    df.groupby(["año", "modalidad"])
    .agg(n=("codigo", "count"),
         monto_total=("valor_referencial_USD", "sum"))
    .reset_index()
)

# Por país y tipo de equipo
by_equipo = (
    adj.groupby(["pais", "tipo_equipo"])
    .agg(n=("codigo", "count"),
         monto_promedio=("valor_referencial_USD", "mean"))
    .reset_index()
    .sort_values(["pais", "n"], ascending=[True, False])
)
by_equipo["monto_promedio_kUSD"] = (by_equipo["monto_promedio"] / 1000).round(0)

# Siemens vs competencia detalle
siemens = adj[adj["marca_ganadora"] == "Siemens Healthineers"]
competencia = adj[adj["marca_ganadora"] != "Siemens Healthineers"]
siemens_detail = (
    siemens.groupby(["pais", "tipo_equipo"])
    .agg(ganadas=("codigo", "count"),
         monto_total=("valor_referencial_USD", "sum"))
    .reset_index()
)

# Tasa de éxito por marca
success = (
    df.groupby("marca_ganadora")
    .agg(participaciones=("codigo", "count"),
         ganadas=("estado", lambda x: (x == "Adjudicado").sum()))
    .reset_index()
)
success = success[success["marca_ganadora"] != ""]
success["tasa_exito_%"] = (success["ganadas"] / success["participaciones"] * 100).round(1)

# ── Crear workbook ──────────────────────────────────────────────────────────
wb = load_workbook("data/licitaciones_mr_ct.xlsx")

def header_style(cell, bg=TEAL, font_color=WHITE, bold=True, size=11):
    cell.font = Font(bold=bold, color=font_color, size=size, name="Calibri")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def data_style(cell, bg=WHITE, bold=False, number_format=None):
    cell.font = Font(bold=bold, color=DARK_GRAY, size=10, name="Calibri")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if number_format:
        cell.number_format = number_format

def alt_row(row_idx):
    return LIGHT_GRAY if row_idx % 2 == 0 else WHITE

thin = Side(style="thin", color=MID_GRAY)
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def write_table(ws, df_in, start_row, start_col, title=None):
    """Escribe un DataFrame como tabla formateada."""
    r = start_row
    c = start_col

    if title:
        ws.merge_cells(
            start_row=r, start_column=c,
            end_row=r, end_column=c + len(df_in.columns) - 1
        )
        tc = ws.cell(r, c, title)
        tc.font = Font(bold=True, color=WHITE, size=12, name="Calibri")
        tc.fill = PatternFill("solid", fgColor=DARK_TEAL)
        tc.alignment = Alignment(horizontal="center", vertical="center")
        for col in range(c, c + len(df_in.columns)):
            ws.cell(r, col).border = border
        ws.row_dimensions[r].height = 28
        r += 1

    # Headers
    for j, col in enumerate(df_in.columns):
        cell = ws.cell(r, c + j, col.replace("_", " ").title())
        header_style(cell)
        cell.border = border
    ws.row_dimensions[r].height = 30
    r += 1

    # Data
    for i, row in df_in.iterrows():
        bg = alt_row(i)
        for j, val in enumerate(row):
            cell = ws.cell(r, c + j, val)
            data_style(cell, bg=bg)
            cell.border = border
            if isinstance(val, float) and val > 1000:
                cell.number_format = '#,##0'
        r += 1

    return r

# ── Hoja: RESUMEN EJECUTIVO ─────────────────────────────────────────────────
if "Resumen_Ejecutivo" in wb.sheetnames:
    del wb["Resumen_Ejecutivo"]
ws_res = wb.create_sheet("Resumen_Ejecutivo", 0)
ws_res.sheet_view.showGridLines = False

# Título principal
ws_res.merge_cells("B2:L3")
tc = ws_res["B2"]
tc.value = "Market Intelligence — Equipos de Imagen Médica (MR & CT)"
tc.font = Font(bold=True, color=WHITE, size=18, name="Calibri")
tc.fill = PatternFill("solid", fgColor=TEAL)
tc.alignment = Alignment(horizontal="center", vertical="center")
ws_res.row_dimensions[2].height = 35
ws_res.row_dimensions[3].height = 35

ws_res.merge_cells("B4:L4")
sub = ws_res["B4"]
sub.value = "Perú (SEACE) & Ecuador (SERCOP)  |  2020–2025  |  Fuente: Datos Abiertos OCDS"
sub.font = Font(italic=True, color=DARK_GRAY, size=11, name="Calibri")
sub.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
sub.alignment = Alignment(horizontal="center", vertical="center")
ws_res.row_dimensions[4].height = 22

# KPI boxes
kpis = [
    ("Total Licitaciones", str(len(df))),
    ("Monto Total (USD)", f"${df['valor_referencial_USD'].sum()/1e6:.1f}M"),
    ("Tasa Adjudicación", f"{len(adj)/len(df)*100:.0f}%"),
    ("Marcas Compitiendo", "4"),
    ("Años Analizados", "6"),
    ("Países", "2"),
]
kpi_cols = [2, 4, 6, 8, 10, 12]
for idx, (label, value) in enumerate(kpis):
    col = kpi_cols[idx]
    ws_res.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col+1)
    ws_res.merge_cells(start_row=7, start_column=col, end_row=7, end_column=col+1)
    ws_res.merge_cells(start_row=8, start_column=col, end_row=8, end_column=col+1)

    lc = ws_res.cell(6, col, label)
    lc.font = Font(bold=False, color=WHITE, size=9, name="Calibri")
    lc.fill = PatternFill("solid", fgColor=DARK_TEAL)
    lc.alignment = Alignment(horizontal="center", vertical="center")

    vc = ws_res.cell(7, col, value)
    vc.font = Font(bold=True, color=WHITE, size=20, name="Calibri")
    vc.fill = PatternFill("solid", fgColor=TEAL)
    vc.alignment = Alignment(horizontal="center", vertical="center")

    ws_res.cell(8, col).fill = PatternFill("solid", fgColor=TEAL)
    ws_res.row_dimensions[6].height = 18
    ws_res.row_dimensions[7].height = 40
    ws_res.row_dimensions[8].height = 10

# Tabla market share
write_table(ws_res, ms_total[["marca_ganadora","licitaciones","share_%","monto_MM_USD"]],
            start_row=10, start_col=2,
            title="Market Share — Total Perú + Ecuador (Adjudicados)")

# Tabla tendencia
write_table(ws_res, tend,
            start_row=10, start_col=9,
            title="Licitaciones por Año y Modalidad")

# Ancho columnas
for col in range(1, 15):
    ws_res.column_dimensions[get_column_letter(col)].width = 16

# ── Hoja: ANÁLISIS COMPETITIVO ──────────────────────────────────────────────
if "Analisis_Competitivo" in wb.sheetnames:
    del wb["Analisis_Competitivo"]
ws_comp = wb.create_sheet("Analisis_Competitivo")
ws_comp.sheet_view.showGridLines = False

ws_comp.merge_cells("B2:M2")
tc2 = ws_comp["B2"]
tc2.value = "Análisis Competitivo — Siemens Healthineers vs. Competencia"
tc2.font = Font(bold=True, color=WHITE, size=14, name="Calibri")
tc2.fill = PatternFill("solid", fgColor=TEAL)
tc2.alignment = Alignment(horizontal="center", vertical="center")
ws_comp.row_dimensions[2].height = 30

write_table(ws_comp, ms_total, start_row=4, start_col=2,
            title="Market Share por Marca")
write_table(ws_comp, success[["marca_ganadora","participaciones","ganadas","tasa_exito_%"]],
            start_row=4, start_col=8,
            title="Tasa de Éxito por Marca")
write_table(ws_comp, by_equipo,
            start_row=16, start_col=2,
            title="Adjudicaciones por País y Tipo de Equipo")

for col in range(1, 15):
    ws_comp.column_dimensions[get_column_letter(col)].width = 18

# ── Hoja: GRÁFICOS ──────────────────────────────────────────────────────────
if "Graficos" in wb.sheetnames:
    del wb["Graficos"]
ws_charts = wb.create_sheet("Graficos")
ws_charts.sheet_view.showGridLines = False

ws_charts.merge_cells("A1:N1")
tc3 = ws_charts["A1"]
tc3.value = "Dashboard Visual — Para importar en Power BI o usar directamente"
tc3.font = Font(bold=True, color=WHITE, size=13, name="Calibri")
tc3.fill = PatternFill("solid", fgColor=DARK_TEAL)
tc3.alignment = Alignment(horizontal="center", vertical="center")
ws_charts.row_dimensions[1].height = 28

# Datos auxiliares para gráficos
# 1. Market share (pie)
start_ms = 3
ws_charts.cell(start_ms, 1, "Marca").font = Font(bold=True, color=WHITE, name="Calibri")
ws_charts.cell(start_ms, 1).fill = PatternFill("solid", fgColor=TEAL)
ws_charts.cell(start_ms, 2, "Licitaciones").font = Font(bold=True, color=WHITE, name="Calibri")
ws_charts.cell(start_ms, 2).fill = PatternFill("solid", fgColor=TEAL)
for i, row in ms_total.iterrows():
    r = start_ms + 1 + list(ms_total.index).index(i)
    ws_charts.cell(r, 1, row["marca_ganadora"])
    ws_charts.cell(r, 2, row["licitaciones"])

# Pie chart market share
pie = PieChart()
pie.title = "Market Share — Licitaciones Ganadas"
pie.style = 10
labels_ref = Reference(ws_charts, min_col=1, min_row=start_ms+1,
                        max_row=start_ms+len(ms_total))
data_ref = Reference(ws_charts, min_col=2, min_row=start_ms,
                     max_row=start_ms+len(ms_total))
pie.add_data(data_ref, titles_from_data=True)
pie.set_categories(labels_ref)
pie.dataLabels = DataLabelList()
pie.dataLabels.showPercent = True
pie.dataLabels.showCatName = True
pie.width = 14
pie.height = 12
ws_charts.add_chart(pie, "D3")

# 2. Tendencia barras
tend_pivot = tend.pivot_table(index="año", columns="modalidad", values="n", fill_value=0).reset_index()
start_tend = 3 + len(ms_total) + 3
ws_charts.cell(start_tend, 1, "Año").font = Font(bold=True, color=WHITE, name="Calibri")
ws_charts.cell(start_tend, 1).fill = PatternFill("solid", fgColor=TEAL)
ws_charts.cell(start_tend, 2, "CT").font = Font(bold=True, color=WHITE, name="Calibri")
ws_charts.cell(start_tend, 2).fill = PatternFill("solid", fgColor=TEAL)
ws_charts.cell(start_tend, 3, "MR").font = Font(bold=True, color=WHITE, name="Calibri")
ws_charts.cell(start_tend, 3).fill = PatternFill("solid", fgColor=TEAL)
for i, row in tend_pivot.iterrows():
    r = start_tend + 1 + i
    ws_charts.cell(r, 1, int(row["año"]))
    ws_charts.cell(r, 2, int(row.get("CT", 0)))
    ws_charts.cell(r, 3, int(row.get("MR", 0)))

bar = BarChart()
bar.type = "col"
bar.grouping = "clustered"
bar.title = "Licitaciones por Año: MR vs CT"
bar.y_axis.title = "N° Licitaciones"
bar.x_axis.title = "Año"
bar.style = 10
data_ref2 = Reference(ws_charts, min_col=2, max_col=3,
                      min_row=start_tend, max_row=start_tend+len(tend_pivot))
cats_ref2 = Reference(ws_charts, min_col=1,
                      min_row=start_tend+1, max_row=start_tend+len(tend_pivot))
bar.add_data(data_ref2, titles_from_data=True)
bar.set_categories(cats_ref2)
bar.shape = 4
bar.width = 14
bar.height = 12
ws_charts.add_chart(bar, "D18")

for col in range(1, 3):
    ws_charts.column_dimensions[get_column_letter(col)].width = 22

# ── Hoja: GUÍA POWER BI ────────────────────────────────────────────────────
if "Guia_PowerBI" in wb.sheetnames:
    del wb["Guia_PowerBI"]
ws_guide = wb.create_sheet("Guia_PowerBI")
ws_guide.sheet_view.showGridLines = False
ws_guide.merge_cells("B2:K2")
gc = ws_guide["B2"]
gc.value = "Guía: Cómo conectar este Excel a Power BI Desktop"
gc.font = Font(bold=True, color=WHITE, size=14, name="Calibri")
gc.fill = PatternFill("solid", fgColor=TEAL)
gc.alignment = Alignment(horizontal="center", vertical="center")
ws_guide.row_dimensions[2].height = 30

steps = [
    ("PASO 1", "Abrir Power BI Desktop → 'Obtener datos' → 'Excel'"),
    ("PASO 2", "Seleccionar este archivo → marcar hoja 'Datos_Licitaciones'"),
    ("PASO 3", "En el editor de Power Query: la columna 'fecha_convocatoria' está en formato fecha ISO — verificar tipo de dato"),
    ("PASO 4", "Cargar datos → Crear los siguientes visuales recomendados:"),
    ("  Visual 1", "Gráfico de barras: Marca ganadora vs N° licitaciones (filtro: País)"),
    ("  Visual 2", "Gráfico circular: Market share % por marca"),
    ("  Visual 3", "Línea de tendencia: Año vs N° licitaciones (separado por MR/CT)"),
    ("  Visual 4", "Mapa: Región vs Monto total (usar lat/lon si se agrega columna)"),
    ("  Visual 5", "Tabla: Top 10 entidades por monto adjudicado"),
    ("  Visual 6", "Tarjetas KPI: Total licitaciones, Monto total, % Siemens"),
    ("PASO 5", "Agregar filtros (slicers): País, Año, Modalidad (MR/CT), Estado"),
    ("PASO 6", "Publicar en Power BI Service → compartir link en portfolio"),
    ("NOTA", "Para datos reales: descargar CSV del SEACE (seace.gob.pe) y reemplazar hoja Datos_Licitaciones manteniendo los mismos nombres de columna"),
]

for i, (paso, desc) in enumerate(steps):
    r = 4 + i
    pc = ws_guide.cell(r, 2, paso)
    pc.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
    pc.fill = PatternFill("solid", fgColor=DARK_TEAL if "PASO" in paso else TEAL)
    pc.alignment = Alignment(horizontal="center", vertical="center")
    pc.border = border

    dc = ws_guide.cell(r, 3, desc)
    dc.font = Font(color=DARK_GRAY, size=10, name="Calibri")
    dc.fill = PatternFill("solid", fgColor=LIGHT_GRAY if i % 2 == 0 else WHITE)
    dc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    dc.border = border
    ws_guide.row_dimensions[r].height = 28

ws_guide.column_dimensions["B"].width = 18
ws_guide.column_dimensions["C"].width = 80
ws_guide.merge_cells(start_row=4, start_column=3, end_row=4, end_column=10)
for r in range(5, 4 + len(steps)):
    ws_guide.merge_cells(start_row=r, start_column=3, end_row=r, end_column=10)

# ── Formato hoja Datos_Licitaciones ────────────────────────────────────────
ws_data = wb["Datos_Licitaciones"]
for cell in ws_data[1]:
    header_style(cell, bg=TEAL)
    cell.border = border
ws_data.auto_filter.ref = ws_data.dimensions
ws_data.freeze_panes = "A2"
for col in ws_data.columns:
    max_len = max(len(str(cell.value or "")) for cell in col)
    ws_data.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

# ── Guardar ─────────────────────────────────────────────────────────────────
out = "data/MR_CT_Market_Analysis_Peru_Ecuador.xlsx"
wb.save(out)
print(f"  Excel profesional guardado: {out}")
print("   Hojas: Resumen_Ejecutivo | Analisis_Competitivo | Graficos | Datos_Licitaciones | Guia_PowerBI")
